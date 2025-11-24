#!/usr/bin/env python3
"""
- Scrapes multiple URLs
- Runs SERP enrichment
- Summarizes chunks for context
- Generates 3 long LinkedIn variants per URL (700–1000 words)
- Streams tokens live while writing
- Saves to Excel + SERP metadata
- Resume-safe with JSON caching
"""

import os, re, time, json
from pathlib import Path
from typing import List, Dict
import requests, validators
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

try:
    from llama_cpp import Llama
except ImportError:
    Llama = None

# ======================================================
# CONFIG
# ======================================================
ROOT = Path(__file__).parent.resolve()
load_dotenv(ROOT / ".env")

SERPAPI_KEY = os.getenv("SERPAPI_KEY")

# Same model path as Script A
MODEL_PATH = ROOT / "Phi-3-mini-4k-instruct-q4.gguf"

EXCEL_PATH = ROOT / "clout_ultra.xlsx"

# Your requested URLs:
BLOG_URLS = [
    "https://leighcuen.substack.com/p/3-trends-reshaping-the-publishing",
    "https://turnerbookwriters.com/blog/trends-in-the-publishing-industry/",
    "https://www.draft2digital.com/blog/industry-trends-with-jane-friedman-ep215/",
    "https://barkerbooks.com/publishing-industry-trends/",
    "https://www.reddit.com/r/writing/comments/1mrfzoi/i_love_writing_but_hate_the-publishing-industry/",
    "https://spines.com/book-publishing-trends-in-2024-authors-should-know/",
    "https://pubrica.com/insights/industry-trends-forecasts-academic-publishing/",
    "https://pagepublishing.com/publishing-and-writing-trends-of-2024/",
    "https://www.netsuite.com/portal/resource/articles/erp/publishing-industry-challenges.shtml"
]

HEADERS = {"User-Agent": "Mozilla/5.0"}

CACHE_SUM = ROOT / "cache_summary.json"
CACHE_VAR = ROOT / "cache_variants.json"
CACHE_SERP = ROOT / "cache_serp.json"

VARIANTS = [
    ("Thought Leadership", "Write a senior thought-leadership article: frameworks, vision, consequences."),
    ("Story Narrative", "Write a narrative article: begin with scene or anecdote, then provoke insight."),
    ("Actionable Framework", "Write an article with strategic takeaways and 3–5 actionable steps."),
]

# ======================================================
# BASIC HELPERS
# ======================================================
def fetch_paras(url: str) -> str:
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except:
        return ""

    soup = BeautifulSoup(r.text, "html.parser")

    # PRIORITY: Typical article containers
    selectors = [
        "article", ".post-content", ".entry-content", ".td-post-content",
        ".elementor-widget-container", ".blog-content", ".content", "#content",
        ".post", ".single-post", ".section-content", ".main-content"
    ]

    # Try each selector; return trimmed text if valid
    for sel in selectors:
        block = soup.select_one(sel)
        if block:
            paras = [p.get_text(" ", strip=True) for p in block.find_all("p")]
            text = "\n\n".join(paras)
            if len(text) > 300:              # skip junk matches
                return text[:15000]          # HARD TRIM (safe for llama)

    # FALLBACK: generic <p> scrape, heavily trimmed
    paras = [p.get_text(" ", strip=True) for p in soup.find_all("p")]
    fallback = "\n\n".join(paras)

    return fallback[:12000]                  # fallback trim


def serp_search(query: str, n: int = 3) -> List[str]:
    if not SERPAPI_KEY:
        return []
    try:
        r = requests.get(
            "https://serpapi.com/search",
            params={"q": query, "api_key": SERPAPI_KEY, "num": n},
            timeout=20).json()
        return [
            o.get("link") for o in r.get("organic_results", [])
            if o.get("link") and validators.url(o.get("link"))
        ][:n]
    except:
        return []

def extract_entities(text: str) -> Dict[str, List[str]]:
    return {
        "emails": sorted(set(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", text))),
        "proper_nouns": sorted(set(re.findall(r"\b[A-Z][a-z]+(?: [A-Z][a-z]+)*\b", text))),
    }

# ======================================================
# LLM
# ======================================================
def load_llm():
    if not MODEL_PATH.exists():
        print("❌ Model missing. Place model file here:\n", MODEL_PATH)
        return None
    print("🔧 Loading model safely…")
    return Llama(
        model_path=str(MODEL_PATH),
        n_ctx=2048,
        n_gpu_layers=20,     # Matches your Mac stability from Script A
        n_threads=os.cpu_count(),
        verbose=False,
    )

def stream_generate(llm, prompt: str, max_toks=450):

    full = ""
    for ch in llm(prompt, max_tokens=max_toks, temperature=0.35, stream=True):
        token = ch.get("choices",[{}])[0].get("text","")
        print(token, end="", flush=True)
        full += token
    return full

# ======================================================
# CACHING
# ======================================================
def load_cache(p): return json.load(open(p)) if p.exists() else {}
def save_cache(p,d): json.dump(d, open(p,"w"), indent=2)

# ======================================================
# CHUNK SUMMARIES
# ======================================================
def summarize_chunks(llm, text):
    c = load_cache(CACHE_SUM)

    chunks = [text[i:i+2500] for i in range(0, len(text), 2500)]

    total = len(chunks)
    results = []

    for i, ch in enumerate(chunks):
        key = f"chunk_{hash(ch) % 99999}"

        # If cached → use it and show skip message
        if key in c:
            summary = c[key]
            word_count = len(summary.split())
            print(f"✔ Cached summary {i+1}/{total} ({word_count} words)")
            results.append(summary)
            continue

        # Show verbose progress message
        print(f"📌 Summarizing chunk {i+1} of {total} ({len(ch.split())} words in source)")

        prompt = f"Summarize clearly, concisely:\n{ch}\nSummary:"
        try:
            r = llm(prompt, max_tokens=40, temperature=0.55)

            out = r['choices'][0]['text'].strip()
        except:
            out = "(failed summary)"

        # Save immediately
        c[key] = out
        save_cache(CACHE_SUM, c)

        results.append(out)

    return "\n".join(results)


# ======================================================
# VARIANT GENERATION
# ======================================================
def generate_variant(llm, label, instr, base, serp):
    c = load_cache(CACHE_VAR)
    key = f"{label}_{hash(base)%99999}"
    if key in c: return c[key]

    prompt=f"""
You are a senior editorial strategist writing a high-impact LinkedIn article.

Variant: {label}
Instruction: {instr}

Write a deep, flowing, human article of 700–1000 words that:
- Expands ideas from **this blog summary**:
{base[:2200]}

- Synthesizes **external insights from other sources**:
{serp[:1500]}

Rules:
- Do NOT summarize; expand, challenge, interpret.
- DO NOT mention “sources” or say “according to…”.
- Use rich paragraphs (5–7 sentences), not bullet lists.
- No academic tone; make it editorial and opinionated.
- Do NOT reveal external scraping.

Format strictly:
###
HEADLINE:
<5–12 word headline>
POST:
<full long-form article>
###
"""
    print(f"\n🎨 Generating {label} …\n")
    text = stream_generate(llm, prompt).strip()

    if "POST:" in text:
        h, b = text.split("POST:",1)
        out = {
            "label": label,
            "headline": h.replace("HEADLINE:","").strip(),
            "body": b.strip()
        }
    else:
        out = {"label":label, "headline":label, "body":text}

    c[key]=out; save_cache(CACHE_VAR,c)
    return out

# ======================================================
# EXCEL
# ======================================================
def init_excel():
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active; ws.title="posts"
        ws.append(["url","variant","headline","body","emails","proper_nouns","serp_links"])
        wb.save(EXCEL_PATH)

def save_row(d):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["posts"]
    ws.append([
        d.get("url"), d.get("variant"),
        d.get("headline"), d.get("body"),
        ", ".join(d.get("emails",[])),
        ", ".join(d.get("proper_nouns",[])),
        ", ".join(d.get("serp_links",[])),
    ])
    wb.save(EXCEL_PATH)

# ======================================================
# MAIN
# ======================================================
def main():
    init_excel()
    llm = load_llm()
    if not llm: return

    for url in BLOG_URLS:
        print(f"\n📝 Scraping blog: {url}")
        blog = fetch_paras(url)
        if not blog:
            print("❌ No readable text. Skipping…")
            continue

        print("🔎 Running SERP search…")
        serp_links = serp_search(blog[:100], n=3)

        serp_text=""
        for s in serp_links:
            serp_text += fetch_paras(s) + "\n\n"; time.sleep(1)

        entities = extract_entities(serp_text)

        print("✂️ Building summary context…")
        summary = summarize_chunks(llm, blog)

        for label, instr in VARIANTS:
            v = generate_variant(llm, label, instr, summary, serp_text)
            save_row({
                "url": url,
                "variant": label,
                "headline": v["headline"],
                "body": v["body"],
                "emails": entities["emails"],
                "proper_nouns": entities["proper_nouns"],
                "serp_links": serp_links
            })
            print(f"\n💾 Saved {label} for {url}")

    print("\n🎉 DONE — Saved to:", EXCEL_PATH)

if __name__=="__main__":
    main()
